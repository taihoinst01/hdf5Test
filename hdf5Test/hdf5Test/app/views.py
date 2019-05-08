"""
Definition of views.
"""

from django.shortcuts import render
from django.http import HttpRequest
from django.http import HttpResponse
from django.http import JsonResponse
from django.template import RequestContext
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt

from .cnn import train
from .cnn import eval

import cv2
import datetime as dt
import h5py
import matplotlib.pyplot as plt
import matplotlib.pylab as plb
import numpy as np
import json
import http.client, urllib.request, urllib.parse, urllib.error, base64
import time
import os
import xlsxwriter
import pandas as pd
import os, sys
from openpyxl import Workbook
from datetime import datetime
from glob import glob
import fnmatch
import requests 
from bs4 import BeautifulSoup
import psycopg2 as pg2 #DB연동
from django.core.files.storage import FileSystemStorage #파일 저장
from openpyxl import load_workbook #excel 읽기
from django.conf import settings
from .module import DomainDicMain
from pandas import DataFrame

DB_CONN_INFO = "host=192.168.0.183 dbname=crawler user=taihoinst password=taiho123 port=5432";

def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)

    try:
        imgForm = "data:image/png;base64,"
        # f = h5py.File('static/h5/data2.h5', 'r')

        # dset = f['X0']
        # data = np.array(dset[:,:,:])
        # cv2.imwrite(file, data)

        # Convert captured image to JPG
        # retval, buffer = cv2.imencode('.png', data)

        # png_as_text = base64.b64encode(buffer)

        # jpg_original = base64.b64decode(png_as_text)

        # imgForm += png_as_text.decode("utf-8") 

        return render(
            request,
            'app/index.html',
            {
                'title':'Home Page',
                'year':datetime.now().year,
                'imgVal':imgForm
            }
        )
    except Exception as e:
        print("[Errno {0}] ".format(e))
        return render(
            request,
            'app/index.html',
            {
                'title':'Home Page',
                'year':datetime.now().year,
                'imgVal':'none'
            }
        )
@csrf_exempt
def getImageFnc(request):
    fileName = request.POST.get('fileName', None)
    fileIndex = int(request.POST.get('fileIndex', None)) +1
    imgForm = "data:image/png;base64, "

    try:
        
        start = dt.datetime.now()

        f = h5py.File('static/h5/data2.h5', 'r')
        
        time.sleep(0.1)

        progTime = dt.datetime.now()
        print("\n1--------------------: ", (progTime - start).seconds, " seconds")

        keys = f.keys()
    
        dset = ''
        fIndex = -1
        fName = ''

        arrLen = len(keys)
        for i, fName in enumerate(keys):
            if fileIndex >= arrLen-1 or fileIndex==0:
                dset = f[fName]
                fIndex = i
                fName = fName
                break
            elif i != 0 and fileIndex <= i:
                dset = f[fName]
                fIndex = i
                fName = fName
                break
        
        
        end = dt.datetime.now()
        print("\n2--------------------: ", (end - start).seconds, " seconds")

        data = np.array(dset[:,:,:])
        # cv2.imwrite(file, data)

        # Convert captured image to JPG
        retval, buffer = cv2.imencode('.png', data)

        png_as_text = base64.b64encode(buffer)

        # jpg_original = base64.b64decode(png_as_text)
    
        imgForm += png_as_text.decode("utf-8") 
        # tmp = str(png_as_text, encoding)

        data = {
            'success': True,
            'imgForm': imgForm,
            'fIndex': fIndex,
            'fName': fName
        }
        return JsonResponse(data)
    except Exception as e:
        print("[Errno {0}] ".format(e))
        data = {
            'success': False,
            'imgForm': '',
            'fIndex': '-1',
            'fName': ''
        }
        return JsonResponse(data)


def dirFileReadFnc(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)

    try:
        imgForm = "data:image/png;base64,"
        # f = h5py.File('static/h5/data2.h5', 'r')

        # dset = f['X0']
        # data = np.array(dset[:,:,:])
        # cv2.imwrite(file, data)

        # Convert captured image to JPG
        # retval, buffer = cv2.imencode('.png', data)

        # png_as_text = base64.b64encode(buffer)

        # jpg_original = base64.b64decode(png_as_text)

        # imgForm += png_as_text.decode("utf-8") 

        return render(
            request,
            'app/normalDir.html',
            {
                'title':'Home Page',
                'year':datetime.now().year,
                'imgVal':imgForm
            }
        )
    except Exception as e:
        print("[Errno {0}] ".format(e))
        return render(
            request,
            'app/normalDir.html',
            {
                'title':'Home Page',
                'year':datetime.now().year,
                'imgVal':'none'
            }
        )


@csrf_exempt
def getImageByDirFnc(request):
    inputName = request.POST.get('fileName', None)
    fileIndex = int(request.POST.get('fileIndex', None))

    try:
        
        start = dt.datetime.now()

        #PATH = os.path.abspath(os.path.join('..', 'input'))
        #C:\Users\taiho\Desktop\images\img_align_celeba
        SOURCE_IMAGES = "C://Users/taiho/Desktop/images/img_align_celeba" #os.path.join(PATH, "images3")

        progTime = dt.datetime.now()
        print("\n1--------------------: ", (progTime - start).seconds, " seconds")
        
        bodyCompare = ''
        saveFileName = ''
        images = []
        fIndex = 0
        for root, dirnames, filenames in os.walk(SOURCE_IMAGES):
            for filename in fnmatch.filter(filenames, '*.*'):
                fullPath = os.path.join(root, filename)
                if inputName == '' and bodyCompare == '':
                    bodyCompare = open(fullPath, 'rb').read()
                    saveFileName = filename
                    fIndex += 1
                    break
                elif fIndex==fileIndex:
                    bodyCompare = open(fullPath, 'rb').read()
                    saveFileName = filename
                    fIndex += 1
                    break
                else:
                    fIndex += 1
                #images.append(os.path.join(root, filename))
                #print(filename)
            if bodyCompare != '':
                print(bodyCompare)
                break
        
        SAVE_IMAGES = "app/static/uploads/" #os.path.join(PATH, "static/app/uploads/")
        saveDir = SAVE_IMAGES + saveFileName
        newFile = open(saveDir, 'wb')
        newFile.write(bodyCompare)
        newFile.close()
        
        end = dt.datetime.now()
        print((end - start).seconds, "seconds", end="")
        data = {
            'success': True,
            'fIndex': fIndex,
            'fName': saveFileName
        }
        return JsonResponse(data)
    except Exception as e:
        print("[Errno {0}] ".format(e))
        data = {
            'success': False,
            'fIndex': -1,
            'fName': ''
        }
        return JsonResponse(data)

def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/contact.html',
        {
            'title':'Contact',
            'message':'Your contact page.',
            'year':datetime.now().year,
        }
    )

def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/about.html',
        {
            'title':'About',
            'message':'Your application description page.',
            'year':datetime.now().year,
        }
    )

def crawlerMLPage(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/crawlerML.html',
        {
            'title':'About',
            'message':'Your application description page.',
            'year':datetime.now().year,
        }
    )
def crawling(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/crawling.html',
        {
            'title':'About',
            'message':'Your application description page.',
            'year':datetime.now().year,
        }
    )

@csrf_exempt
def webcrawlerStart(request):
        """Renders the about page."""
        assert isinstance(request, HttpRequest)
        resultlist = webcrawlerApp();
        data = {
                'success': True,
                'result' : resultlist
            }
        return JsonResponse(data)

def webcrawlerApp():
    cnt = 0
    pageNum = 100
    resultlist = []
    queryList = []
    num = 0
    while cnt<=pageNum:
       r = requests.get('http://minishop.gmarket.co.kr/cam365/List?Title=Best%20Item&CategoryType=General&SortType=MostPopular&DisplayType=List&Page='+str(cnt)+'&PageSize=60&IsFreeShipping=False&HasDiscount=False&HasStamp=False&HasMileage=False&IsInternationalShipping=False&MinPrice=36890&MaxPrice=912120#listTop')
       html = r.text
       soup = BeautifulSoup(html, 'html.parser')
       titles = soup.select('.sbj') 
       for title in titles:
            resultlist.append(title.text)
            queryList.append('INSERT INTO public.\"TBL_CRAWLER_RESULT_LIST\"(\"SENTENCE\") VALUES (\'' + title.text.split('\n')[1] + '\');')
       cnt += 1

    result = dbInsertQuery(queryList)
    return resultlist
    
def getCrawlerResultListFnc(request):
    try :
        conn = pg2.connect(DB_CONN_INFO)
        cur = conn.cursor()

        cur.execute('SELECT * FROM public."TBL_CRAWLER_RESULT_LIST";')
        rows = cur.fetchall()

        data = {
            'success': True,
            'crawlerResultList': rows,
        }

        return JsonResponse(data)
    except Exception as e:
        print('postgresql database connection error!')
        print(e)
    finally:
        if conn:
            conn.close()

def uploadExcelFnc(request):
    if request.method == 'POST' and request.FILES['excel']:
        excel = request.FILES['excel']
        fs = FileSystemStorage()
        filename = fs.save(excel.name, excel)
        worksheet = load_workbook('c:/pythonFiles' + '\\' + filename, data_only=True).worksheets[0]
        queryList = []
        for column in worksheet.iter_rows(min_row=2):
            sentence = str(column[4].value)
            if column[0] is None:
                break
            queryList.append('INSERT INTO public.\"TBL_CRAWLER_RESULT_LIST\"(\"SENTENCE\") VALUES (\'' + sentence + '\');')

        result = dbInsertQuery(queryList)

        data = {
            'success': result
        }
    return JsonResponse(data)


@csrf_exempt
def mlexcelexport(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    isDone = False
    try:
        mlexcelexportFnc()
    except Exception as e:
        print(e)
    else:
        isDone = True
    finally:
        data = {
                'success': isDone
            }
        return JsonResponse(data)

def mlexcelexportFnc():

    query = 'SELECT * FROM public."TBL_CRAWLER_RESULT_LIST";'
    result = dbSelectQuery(query)

    dt = datetime.now()
       
    list_values = [[result[0][0]]]
    

    # Create a Workbook on Excel:
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'data'

    # Print the titles into Excel Workbook:
    row = 1
    sheet['A'+str(row)] = 'SEQ'
    sheet['B'+str(row)] = 'Sentence'
    sheet['C'+str(row)] = 'Company Name'
    sheet['D'+str(row)] = '정확도'
    sheet['E'+str(row)] = 'Model Numbeer'
    sheet['F'+str(row)] = '정확도'

    # Populate with data
    for item in list_values:
        row += 1
        sheet['A'+str(row)] = item[0]
        sheet['B'+str(row)] = item[1]
        sheet['C'+str(row)] = item[2]

    # Save a file by date:
    filename = 'data_' + dt.strftime("%Y%m%d_%I%M%S") + '.xlsx'
    wb.save(filename)

    # Open the file for the user:
    os.chdir('c:/pythonFiles')
    



def dbInsertQuery(queryList):
    try :
        conn = pg2.connect(DB_CONN_INFO)
        cur = conn.cursor()

        for query in queryList:
            cur.execute(query)
        conn.commit()
        print ("Record inserted successfully")

    except Exception as e:
        print(e)
        return False
    else:
        return True
    finally:
        if conn:
            conn.close()

@csrf_exempt
def mlProcessFnc(request):
    data = {
                'success': False
            }
    try:
        query = 'SELECT "SENTENCE" FROM public."TBL_CRAWLER_RESULT_LIST";'
        result = dbSelectQuery(query)
        modelNumberList = []
        if(result):
            #for row in result:
                #modelNumber = DomainDicMain.run(row[0])
                #modelNumberList.append(modelNumber)
        
            data_file_path__ = 'app/cnn/data/kkk.cls'
            col_file_Path__ = 'app/cnn/data/kkk.train'
            
            #with open(data_file_path__, 'rb') as data:
            #    col_list = [l.decode('utf8', 'ignore') for l in data.readlines()]
        
            #with open(col_file_Path__, 'rb') as data:
            #    train_list = [l.decode('utf8', 'ignore') for l in data.readlines()]
                
            #file = open(col_file_Path__, 'a', -1, encoding='UTF8')
            #file_cls = open(data_file_path__, 'a', -1, encoding='UTF8')
            #new_col_list = []
            #new_val_list = False
            #for item in result:
            #    rep_item = item[0].strip().replace(",","")
            #    if any(item[0] in s for s in train_list):
            #        continue
            #    else:
            #        is_in = False
            #        col_val = ""
            #        for cvalue in col_list:
            #            cvalue = cvalue.replace("\n","").replace("\r","")
            #            if cvalue in rep_item:
            #                is_in = True
            #                col_val = cvalue

            #        if is_in:
            #            file.write("\n" + rep_item + "," + col_val)
            #            train_list.append(rep_item + "," + col_val)
            #            new_val_list = True
            #        else:
                        
            #            if "HANKOOK" in rep_item:
            #                file.write("\n" + rep_item + ",한국타이어")
            #                new_col_list.append("hankook")
            #                col_list.append("hankook")
            #                continue

            #            endIndex = -1
            #            if rep_item[0] == '[':
            #                for idx, str in enumerate(item[0]):
            #                    if str == ']':
            #                        endIndex = idx
            #                        break
                        
            #            if endIndex==-1:
            #                file.write("\n" + rep_item + ",etc")
            #            else:
            #                companyName = rep_item[1:endIndex]
            #                if companyName not in new_col_list:
            #                    file.write("\n" + rep_item + "," + companyName)
            #                    file_cls.write("\n" + companyName)
            #                    new_col_list.append(companyName)
            #                    col_list.append(companyName)
            #                #train_list.append(item[0] + ",etc" + "\n")
            #file.close()
            #file_cls.close()

            #eval_list = []
            #for fil in train_list:
            #    obDict = {}
            #    obDict['text'] = fil.replace("\n", "").split(",")[0]
            #    obDict['location'] = fil.replace("\n", "").split(",")[1]
            #    eval_list.append(obDict)

            #if (len(new_col_list)>0 or new_val_list):
            #    train.startTrain()
            #train.startTrain()
            #rst_list = eval.startEval(eval_list)
            eval_list = []
            for item in result:
                rep_item = item[0].strip().replace(",","")
                obDict = {}
                obDict['text'] = rep_item
                eval_list.append(obDict)

            rst_list = eval.startEval(eval_list)

            data = {
                'success': True,
                'modelNumberList': modelNumberList, 
                'rst_list' : rst_list
            }
        else:
            print(e)
    except Exception as e:
        print(e)
    finally:
        return JsonResponse(data)
    
def dbSelectQuery(query):
    try :
        conn = pg2.connect(DB_CONN_INFO)
        cur = conn.cursor()

        cur.execute(query)
        rows = cur.fetchall()

    except Exception as e:
        print(e)
        return False
    else:
        return rows
    finally:
        if conn:
            conn.close()